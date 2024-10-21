from flask import Flask, render_template, request, redirect, url_for, flash, jsonify
from models import Base, User, Pdf
from sqlalchemy.orm import sessionmaker
from sqlalchemy import create_engine
from flask import session as login_session
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from urllib.request import urlopen
from pdfminer.pdfparser import PDFParser
from pdfminer.pdfdocument import PDFDocument
import requests, json, os, weasyprint


engine = create_engine('mysql+mysqlconnector://root@localhost/web_scraping')
Base.metadata.bind = engine
DBSession = sessionmaker(bind = engine)
session = DBSession()
app = Flask(__name__)


@app.route("/", methods=["GET", "POST"])
def login():
    if login_session.get("email") is not None:
        flash("Your account is already logged in.")
        return redirect(url_for("url"))
    else:
        if request.method == "POST":
            try:
                user = session.query(User).filter_by(email=request.form["email"], password=request.form["password"]).one()
            except:
                flash("Invalid input.")
                return redirect(url_for("login"))


            login_session["id"] = user.id
            login_session["email"] = user.email
            return redirect(url_for("url"))
        else:
            return render_template("login.html")


@app.route("/logout")
def logout():
    login_session.clear()
    flash("Logged out successfully.")
    return redirect(url_for("login"))


@app.route("/register", methods=["GET", "POST"])
def register():
    if request.method == "POST":
        try:
            user = session.query(User).filter_by(email=request.form["email"]).one()
            flash("User already registered or email already taken.")
        except:
            user = User(name=request.form["name"], email=request.form["email"], password=request.form["password"])
            session.add(user)
            session.commit()
            flash("User registered successfully.")


        return redirect(url_for("login"))
    else:
        return render_template("register.html")


@app.route("/url", methods=["GET", "POST"])
def url():
    if login_session.get("email") is None:
        flash("You must log in first.")
        return redirect(url_for("login"))
    else:
        if request.method == "POST":
            url = "https://" + request.form["url"]
            if request.form["url"] == "cyberleninka.org":
                url = url + "/article/c/computer-and-information-sciences"
                response = requests.get(url)
                url = url.replace("/article/c/computer-and-information-sciences", "")


                soup = BeautifulSoup(response.content, "html.parser")
                links = soup.find("div", class_="full").find_all("a")
                print_urls = [url + link["href"] + ".pdf" for link in links]
                return render_template("url.html", articles = print_urls)


            if request.form["url"] == "hanushek.stanford.edu":
                url = url + "/publications/academic"
                response = requests.get(url)
                soup = BeautifulSoup(response.content, "html.parser")
                links = soup.find_all("div", class_="views-field views-field-field-publications-download")
                print_urls = []


                for link in links:
                    a = link.find("a")
                    try:
                        print_urls.append(a["href"])
                    except:
                        pass


                return render_template("url.html", articles = print_urls)


            if request.form["url"] == "eric.ed.gov":
                url = url + "/?q=source%3A%22Academic+Leadership+Journal+in+Student+Research%22"
                response = requests.get(url)
                soup = BeautifulSoup(response.content, "html.parser")
                links = soup.find_all("div", class_="r_i")


                print_urls = [link.find("div", class_="r_f").find("a")["href"] for link in links]


                return render_template("url.html", articles = print_urls)


            if request.form["url"] == "ecommons.aku.edu":
                url = url + "/pjns/"
                response = requests.get(url)
                soup = BeautifulSoup(response.content, "html.parser")
                links = soup.find_all("p", class_="pdf")


                print_urls = [link.find("a")["href"] for link in links]


                return render_template("url.html", articles = print_urls)


            # if request.form["url"] == "www.ezinearticles.com":
            #     url = "https://www.ezinearticles.com"
            #     response = requests.get(url)


            #     soup = BeautifulSoup(response.content, "html.parser")
            #     articleid = soup.find(id='recent-article-container')
            #     articles = articleid.find_all("div", class_="article")


            #     links = [article.find("h3").find("a")['href'] for article in articles]


            #     full_urls = [url + link for link in links]


            #     return render_template("url.html", articles = full_urls)


            else:
                flash("Article not found.")
                return redirect(url_for("url"))
        else:
            return render_template("url.html")


@app.route("/download", methods=["POST"])
def download():
    if login_session.get("email") is None:
        flash("You must log in first.")
        return redirect(url_for("login"))
    else:
        pdf = str(request.form["url"])
        if "https://cyberleninka.org" in pdf: 
            pdf = pdf.replace("https://cyberleninka.org/article/n/", "")
            pdf_file = "./static/pdfs/" + pdf
            try:
                pdf_db = session.query(Pdf).filter_by(name = pdf_file, user_id = login_session["id"]).one()
                flash("PDF already downloaded.")
            except:
                response = urlopen(request.form["url"])
                file = open(pdf_file, "wb")
                file.write(response.read())
                file.close()
                pdf_db = Pdf(name = pdf_file, user_id = login_session["id"])
                session.add(pdf_db)
                session.commit()
                flash("PDF downloaded successfully.")


            return redirect(url_for("url"))


        if "http://hanushek.stanford.edu" in pdf:
            pdf = pdf.replace("http://hanushek.stanford.edu/sites/default/files/publications/", "")
            pdf_file = "./static/pdfs/" + pdf
            try:
                pdf_db = session.query(Pdf).filter_by(name = pdf_file, user_id = login_session["id"]).one()
                flash("PDF already downloaded.")
            except:
                response = urlopen(request.form["url"])
                file = open(pdf_file, "wb")
                file.write(response.read())
                file.close()
                pdf_db = Pdf(name = pdf_file, user_id = login_session["id"])
                session.add(pdf_db)
                session.commit()
                flash("PDF downloaded successfully.")


            return redirect(url_for("url"))


        # if "https://www.ezinearticles.com" in pdf:
        #     pdf = pdf.replace("https://www.ezinearticles.com/?", "")
        #     pdf_file = "./static/pdfs/" + pdf + ".pdf"
        #     try:
        #         pdf_db = session.query(Pdf).filter_by(name = pdf_file, user_id = login_session["id"]).one()
        #         flash("PDF already downloaded.")
        #     except:
        #         weasyprint.HTML(request.form["url"]).write_pdf(pdf_file)
        #         pdf_db = Pdf(name = pdf_file, user_id = login_session["id"])
        #         session.add(pdf_db)
        #         session.commit()
        #         flash("PDF downloaded successfully.")


        #     return redirect(url_for("url"))


        if "http://files.eric.ed.gov" in pdf:
            pdf = pdf.replace("http://files.eric.ed.gov/fulltext/", "")
            pdf_file = "./static/pdfs/" + pdf
            try:
                pdf_db = session.query(Pdf).filter_by(name = pdf_file, user_id = login_session["id"]).one()
                flash("PDF already downloaded.")
            except:
                response = urlopen(request.form["url"])
                file = open(pdf_file, "wb")
                file.write(response.read())
                file.close()
                pdf_db = Pdf(name = pdf_file, user_id = login_session["id"])
                session.add(pdf_db)
                session.commit()
                flash("PDF downloaded successfully.")


            return redirect(url_for("url"))


        if "https://ecommons.aku.edu" in pdf:
            pdf = pdf.replace("https://ecommons.aku.edu/cgi/viewcontent.cgi?", "")
            pdf = pdf.replace("&context=pjns", "")
            pdf = pdf.replace("=", "")
            pdf_file = "./static/pdfs/" + pdf
            try:
                pdf_db = session.query(Pdf).filter_by(name = pdf_file, user_id = login_session["id"]).one()
                flash("PDF already downloaded.")
            except:
                response = urlopen(request.form["url"])
                file = open(pdf_file, "wb")
                file.write(response.read())
                file.close()
                pdf_db = Pdf(name = pdf_file, user_id = login_session["id"])
                session.add(pdf_db)
                session.commit()
                flash("PDF downloaded successfully.")


            return redirect(url_for("url"))

            
        else:
            flash("There's something wrong with the download.")
            return redirect(url_for("url"))


@app.route("/pdfs")
def pdfs():
    if login_session.get("email") is None:
        flash("You must log in first.")
        return redirect(url_for("login"))
    else:
        try:
            pdf = session.query(Pdf).filter_by(user_id = login_session["id"]).all()
            flash("Here is the result.")
            return render_template("pdfs.html", articles = pdf)
        except:
            flash("No article downloaded.")
            return render_template("pdfs.html")


@app.route("/excel")
def excel():
    if login_session.get("email") is None:
        flash("You must log in first.")
        return redirect(url_for("login"))
    else:
        pdfs = session.query(Pdf).filter_by(user_id = login_session["id"]).all()
        if not pdfs:
            flash("Can't be converted to excel.")
            if os.path.exists(login_session["email"]+".xlsx"):
                os.remove(login_session["email"]+".xlsx")
            return redirect(url_for("pdfs"))


        myresult = []
        title = ('Title', 'CreationDate', 'Creator', 'ModDate', 'Producer')
        myresult.append(title)
        for pdf in pdfs:
            fp = open(pdf.name, 'rb')
            parser = PDFParser(fp)
            doc = PDFDocument(parser)
            result = (doc.info[0].get("Title").decode('UTF-8'), doc.info[0].get("CreationDate").decode('UTF-8'),\
                doc.info[0].get("Creator").decode('UTF-8'), doc.info[0].get("ModDate").decode('UTF-8')\
                    , doc.info[0].get("Producer").decode('UTF-8'))
            myresult.append(result)


        wb = Workbook()
        sheet = wb.active
        sheet.title = 'PDF Report'
        sheet.cell(column=1, row=1).value="PDF Report List"
        for row in range(1, len(myresult)+1):
            for col in range(1, len(myresult[row-1])+1):
                sheet.cell(column=col, row=row+1).value = myresult[row-1][col-1]
        wb.save(login_session["email"]+'.xlsx')
        flash("Successfully converted to excel.")
        return redirect(url_for("pdfs"))


@app.route("/jsons")
def jsons():
    if login_session.get("email") is None:
        flash("You must log in first.")
        return redirect(url_for("login"))
    else:
        pdfs = session.query(Pdf).filter_by(user_id = login_session["id"]).all()
        if not pdfs:
            flash("Information is empty.")
            if os.path.exists(login_session["email"]+".json"):
                os.remove(login_session["email"]+".json")
            return redirect(url_for("pdfs"))


        try:
            wb_obj = load_workbook(login_session["email"] + ".xlsx")
            sheet_obj = wb_obj.active
            max_col = sheet_obj.max_column
            max_row = sheet_obj.max_row
            item = []
            items = []


            for i in range(3, max_row+1):
                for j in range(1, max_col+1):
                    cell_obj = sheet_obj.cell(row = i, column = j)
                    item.append(cell_obj.value)

                
                temp = item[:]
                items.append(temp)
                item.clear()
        except:
            flash("You must generate the excel file.")
            return redirect(url_for("pdfs"))


        return jsonify([serialize(item) for item in items])
        # json_object = json.dumps([serialize(item) for item in items], indent=4)
        # file = open(login_session["email"]+".json", "w")
        # file.write(json_object)
        # file.close()
        # flash("JSON file successfully created.")
        # return redirect(url_for("pdfs"))


def serialize(elem):
    return{
        "Title":elem[0],
        "CreationDate":elem[1],
        "Creator":elem[2],
        "ModDate":elem[3],
        "Producer":elem[4]
    }


@app.route("/deletepdf", methods=["POST"])
def deletepdf():
    if login_session.get("email") is None:
        flash("You must log in first.")
        return redirect(url_for("login"))
    else:
        try:
            pdf = session.query(Pdf).filter_by(name = request.form["deletepdf"], user_id = login_session["id"]).one()
            session.delete(pdf)
            session.commit()
            flash("PDF successfully deleted.")
            return redirect(url_for("pdfs"))
        except:
            flash("Error on deletion.")
            return redirect(url_for("pdfs"))
        

if __name__ == "__main__":
    app.secret_key = "mohsin"
    app.debug = True
    app.run(host = "127.0.0.1", port = 8080)